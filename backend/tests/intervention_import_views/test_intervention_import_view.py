import json
import tempfile
from pathlib import Path
from unittest.mock import patch

import mongomock
import openpyxl
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client

from core.models import Intervention
from core.views.intervention_import import (
    _guess_file_media_type,
    _guess_media_type_for_url,
    _guess_provider,
    _is_raw_file_name,
    _map_content_type,
    _normalize_lang,
    _parse_bool,
    _parse_duration_minutes,
    _parse_external_id_and_language,
    _parse_int,
    _split_list,
    _spotify_embed,
    _validate_id_format,
    _youtube_embed,
    import_interventions_from_excel,
)

client = Client()


@pytest.fixture(autouse=True, scope="function")
def mongo_mock():
    from mongoengine import connect, disconnect
    from mongoengine.connection import _connections

    alias = "default"
    if alias in _connections:
        disconnect(alias)

    conn = connect(
        "mongoenginetest",
        alias=alias,
        host="mongodb://localhost",
        mongo_client_class=mongomock.MongoClient,
    )
    yield conn
    disconnect(alias)


def test_helper_parsers():
    # Legacy 2-part format
    assert _parse_external_id_and_language("4001_de") == ("4001", "de", None)
    # New 3-part format: {number}_{format}_{lang}
    assert _parse_external_id_and_language("3500_web_de") == ("3500_web", "de", "web")
    assert _parse_external_id_and_language("30500_vid_pt") == ("30500_vid", "pt", "vid")
    assert _parse_external_id_and_language("3500_br_it") == ("3500_br", "it", "br")
    # Unknown format code → format_code is None, but external_id still correct
    assert _parse_external_id_and_language("3500_xyz_de") == ("3500_xyz", "de", None)
    assert _map_content_type("video") == "Video"
    assert _map_content_type("vid") == "Video"
    assert _map_content_type("br") == "Brochure"
    assert _map_content_type("gfx") == "Graphics"
    assert _parse_duration_minutes("10-20") == 15
    assert _split_list("a, b; a") == ["a", "b"]


def test_helper_bool_int_lang_provider_and_media():
    assert _parse_bool("yes") is True
    assert _parse_bool("no", True) is False
    assert _parse_bool("??", True) is True
    assert _parse_int("22") == 22
    assert _parse_int("x") is None

    assert _normalize_lang("german") == "de"
    assert _normalize_lang("italiano") == "it"
    assert _normalize_lang("português") == "pt"
    assert _normalize_lang("nl") == "nl"
    assert _normalize_lang("dutch") == "nl"
    assert _normalize_lang("unknown") is None

    assert _guess_provider("https://open.spotify.com/track/abc") == "spotify"
    assert _guess_provider("https://youtu.be/abcd1234") == "youtube"
    assert _guess_provider("https://example.com") == "website"
    assert _guess_provider("") is None

    assert _spotify_embed("https://open.spotify.com/track/abc123")
    assert _spotify_embed("https://example.com") is None
    assert _youtube_embed("https://youtu.be/abcd1234")
    assert _youtube_embed("https://example.com") is None

    assert _guess_media_type_for_url("https://example.com/a.mp3", "website") == "audio"
    assert _guess_media_type_for_url("https://example.com/a.pdf", "website") == "pdf"
    assert _guess_media_type_for_url("", "website") == "website"
    assert _guess_file_media_type("clip.mp4", "text") == "video"
    assert _guess_file_media_type("image.png", "text") == "image"
    assert _is_raw_file_name("file.pdf") is True
    assert _is_raw_file_name("https://example.com/file.pdf") is False


def test_validate_id_format():
    # Valid new 3-part IDs produce no warnings
    assert _validate_id_format("3500_web_de") == []
    assert _validate_id_format("30500_vid_pt") == []
    assert _validate_id_format("3500_br_it") == []
    # Legacy 2-part ID — no warnings (backward compat)
    assert _validate_id_format("4001_de") == []
    # Unknown format code
    msgs = _validate_id_format("3500_xyz_de")
    assert any("xyz" in m for m in msgs)
    # Invalid prefix length
    msgs = _validate_id_format("350_web_de")
    assert any("350" in m for m in msgs)
    # Unknown language
    msgs = _validate_id_format("3500_web_xx")
    assert any("xx" in m for m in msgs)


def test_more_helper_branches_for_import_module():
    assert _map_content_type("app") == "App"
    assert _map_content_type("website") == "Website"
    assert _map_content_type("image") == "Graphics"
    assert _map_content_type("podcast") == "Audio"
    assert _map_content_type("brochure") == "Brochure"
    assert _map_content_type("") == "Brochure"

    assert _parse_duration_minutes("< 10") == 10
    assert _parse_duration_minutes("about 25 min") == 25
    assert _parse_duration_minutes("") is None

    assert _normalize_lang("français") == "fr"
    assert _normalize_lang("eng") == "en"
    assert _normalize_lang("") is None

    assert _guess_provider("https://soundcloud.com/x") == "soundcloud"
    assert _guess_provider("https://vimeo.com/123") == "vimeo"
    assert _guess_media_type_for_url("https://example.com/x.webp", "website") == "image"
    assert _guess_media_type_for_url("https://example.com/x.mp4", "website") == "video"
    assert _guess_media_type_for_url("https://example.com", "app") == "app"
    assert _guess_file_media_type("track.ogg", "text") == "audio"
    assert _guess_file_media_type("file.unknown", "text") == "text"


def test_import_endpoint_method_and_validation():
    r1 = client.get("/api/interventions/import/excel", HTTP_AUTHORIZATION="Bearer test")
    assert r1.status_code == 405

    r2 = client.post("/api/interventions/import/excel", data={}, HTTP_AUTHORIZATION="Bearer test")
    assert r2.status_code == 400

    bad = SimpleUploadedFile("bad.txt", b"x", content_type="text/plain")
    r3 = client.post(
        "/api/interventions/import/excel",
        data={"file": bad},
        HTTP_AUTHORIZATION="Bearer test",
    )
    assert r3.status_code == 400


@patch(
    "core.views.intervention_import.import_interventions_from_excel",
    return_value={"created": 1, "updated": 0, "skipped": 0, "errors": []},
)
def test_import_endpoint_success_with_mocked_service(mock_import):
    f = SimpleUploadedFile(
        "interventions.xlsx",
        b"dummy",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    r = client.post(
        "/api/interventions/import/excel",
        data={"file": f},
        HTTP_AUTHORIZATION="Bearer test",
    )
    assert r.status_code == 200
    assert r.json()["success"] is True
    assert r.json()["created"] == 1


def test_import_interventions_from_excel_dry_run_and_save():
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "x.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Content"
        ws.append(
            [
                "intervention_id",
                "title",
                "description",
                "content_type",
                "link",
                "where (multi-choice)",
                "setting (multi-choice)",
                "duration",
            ]
        )
        ws.append(
            [
                "4001_de",
                "T1",
                "D1",
                "Video",
                "https://example.com/v1",
                "outside",
                "individual",
                "10-20",
            ]
        )
        wb.save(p)

        dry = import_interventions_from_excel(str(p), dry_run=True)
        assert dry["created"] == 1
        assert Intervention.objects.count() == 0

        real = import_interventions_from_excel(str(p), dry_run=False)
        assert real["created"] == 1
        assert Intervention.objects.count() == 1
        doc = Intervention.objects.first()
        assert doc.external_id == "4001"
        assert doc.language == "de"
        assert doc.content_type == "Video"
        assert doc.where == ["outside"]
        assert doc.setting == ["individual"]
        assert len(doc.media) == 1


def test_import_interventions_from_excel_missing_required_columns_raises():
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "x.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Content"
        ws.append(["intervention_id", "title"])  # intentionally incomplete
        ws.append(["4001_de", "T1"])
        wb.save(p)

        with pytest.raises(ValueError):
            import_interventions_from_excel(str(p), dry_run=True)


def test_import_interventions_from_excel_header_fallback_and_update_dedupe():
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / "x.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Content (2)"  # tests sheet fallback by prefix
        ws.append(
            [
                "id",
                "title",
                "description",
                "format",
                "link (text input)",
                "where (multi-choice)",
                "setting (multi-choice)",
                "topic",
                "keywords",
                "language",
            ]
        )
        ws.append(
            [
                "5001_en",
                "T1",
                "D1",
                "Web",
                "https://example.com/item",
                "home\noutside",
                "individual|group",
                "balance|mobility",
                "tag1\ntag2",
                "english",
            ]
        )
        wb.save(p)

        first = import_interventions_from_excel(str(p), sheet_name="Content", dry_run=False)
        assert first["created"] == 1

        # second pass should update same doc, not create duplicate
        second = import_interventions_from_excel(str(p), sheet_name="Content", dry_run=False)
        assert second["updated"] == 1
        assert Intervention.objects(external_id="5001", language="en").count() == 1

        doc = Intervention.objects(external_id="5001", language="en").first()
        assert doc is not None
        assert "home" in doc.where and "outside" in doc.where
        assert "individual" in doc.setting and "group" in doc.setting
        assert len(doc.media or []) >= 1
