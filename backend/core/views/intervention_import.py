# core/services/intervention_import.py

import os
import re
import tempfile
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from mongoengine.queryset.visitor import Q
from rest_framework.decorators import permission_classes
from rest_framework.permissions import IsAuthenticated

from core.models import Intervention, InterventionMedia

# ---------------- view helpers ----------------


def _parse_bool(v: Any, default: bool = False) -> bool:
    if v is None:
        return default
    s = str(v).strip().lower()
    if s in ("1", "true", "yes", "y", "on"):
        return True
    if s in ("0", "false", "no", "n", "off"):
        return False
    return default


def _parse_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(str(v).strip())
    except Exception:
        return None


def _bad(message: str, status: int = 400, **extra: Any) -> JsonResponse:
    payload: Dict[str, Any] = {"success": False, "error": message}
    payload.update(extra)
    return JsonResponse(payload, status=status)


# ---------------- import endpoint ----------------


@csrf_exempt
@permission_classes([IsAuthenticated])
def import_interventions(request):
    """
    POST /api/interventions/import/
    multipart/form-data

    Required:
      - file: .xlsx or .xlsm

    Optional:
      - sheet_name (default "Content")
      - dry_run ("true"/"false") default false
      - limit (int) optional
      - default_lang (default "en")

    Returns:
      { created, updated, skipped, errors }
    """
    print(
        "Received intervention import request:",
        request.POST.dict(),
        "Files:",
        request.FILES,
    )
    if request.method != "POST":
        return _bad("Method not allowed", status=405)

    up = request.FILES.get("file")
    if not up:
        return _bad("Missing file. Please upload an .xlsx or .xlsm file.", status=400)

    filename = (getattr(up, "name", "") or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xlsm")):
        return _bad("Invalid file type. Only .xlsx or .xlsm are allowed.", status=400)

    sheet_name = (request.POST.get("sheet_name") or "Content").strip() or "Content"
    dry_run = _parse_bool(request.POST.get("dry_run"), False)
    default_lang = (request.POST.get("default_lang") or "en").strip().lower() or "en"
    limit = _parse_int(request.POST.get("limit"))

    tmp_path: Optional[str] = None

    try:
        suffix = ".xlsm" if filename.endswith(".xlsm") else ".xlsx"

        # Save upload to disk (openpyxl reads filesystem path nicely)
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            for chunk in up.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        result = import_interventions_from_excel(
            xlsm_path=tmp_path,
            sheet_name=sheet_name,
            dry_run=dry_run,
            limit=limit,
            default_lang=default_lang,
        )

        return JsonResponse(
            {
                "success": True,
                **(result or {}),
            },
            status=200,
        )

    except Exception as e:
        return _bad("Import failed.", status=500, details=str(e))

    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


# ---------------- Excel parsing helpers ----------------


def _norm(s: Any) -> str:
    return (str(s).strip() if s is not None else "").strip()


def _split_list(s: str) -> List[str]:
    s = _norm(s)
    if not s:
        return []
    parts = re.split(r"[;,]\s*", s)
    out: List[str] = []
    for p in parts:
        p = _norm(p)
        if p:
            out.append(p)
    # de-dup preserving order
    seen = set()
    uniq: List[str] = []
    for x in out:
        k = x.lower()
        if k not in seen:
            uniq.append(x)
            seen.add(k)
    return uniq


def _parse_duration_minutes(s: str) -> Optional[int]:
    s = _norm(s).lower()
    if not s:
        return None
    m = re.search(r"(\d+)\s*-\s*(\d+)", s)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        return int(round((a + b) / 2))
    m = re.search(r"<\s*(\d+)", s)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d+)", s)
    if m:
        return int(m.group(1))
    return None


def _map_content_type(x: str) -> str:
    s = _norm(x).lower()
    if not s:
        return "Text"
    if "video" in s:
        return "Video"
    if "audio" in s or "podcast" in s or "sound" in s:
        return "Audio"
    if "image" in s or "picture" in s:
        return "Image"
    if "app" in s:
        return "App"
    if "web" in s or "site" in s or "link" in s or "website" in s:
        return "Website"
    if "pdf" in s or "text" in s or "brochure" in s:
        return "Text"
    return "Text"


def _parse_external_id_and_language(intervention_id: str) -> Tuple[str, Optional[str]]:
    """
    Excel uses e.g. '4001_de', '4001_it'. We want:
      external_id='4001', language='de'
    """
    s = _norm(intervention_id)
    m = re.match(r"^(.+?)_([a-z]{2})$", s.lower())
    if m:
        return m.group(1), m.group(2)
    return s, None


def _normalize_lang(raw: Any) -> Optional[str]:
    s = _norm(raw).lower()
    if not s:
        return None
    if s in ("de", "ger", "deu", "german", "deutsch"):
        return "de"
    if s in ("en", "eng", "english"):
        return "en"
    if s in ("fr", "fra", "fre", "french", "français", "francais"):
        return "fr"
    if s in ("it", "ita", "italian", "italiano"):
        return "it"
    return None


def _guess_provider(url: str) -> Optional[str]:
    u = _norm(url).lower()
    if not u:
        return None
    if "spotify.com" in u:
        return "spotify"
    if "youtu.be" in u or "youtube.com" in u:
        return "youtube"
    if "soundcloud.com" in u:
        return "soundcloud"
    if "vimeo.com" in u:
        return "vimeo"
    return "website"


def _spotify_embed(url: str) -> Optional[str]:
    u = _norm(url)
    m = re.search(r"open\.spotify\.com/(track|playlist|album|episode|show)/([A-Za-z0-9]+)", u)
    if not m:
        return None
    typ, sid = m.group(1), m.group(2)
    return f"https://open.spotify.com/embed/{typ}/{sid}"


def _youtube_embed(url: str) -> Optional[str]:
    u = _norm(url)
    m = re.search(r"(?:youtu\.be/|v=)([A-Za-z0-9_-]{6,})", u)
    if not m:
        return None
    vid = m.group(1)
    return f"https://www.youtube.com/embed/{vid}"


def _guess_media_type_for_url(url: str, content_type_fallback: str) -> str:
    u = _norm(url).lower()
    if not u:
        return content_type_fallback

    if "spotify.com" in u:
        return "streaming"
    if "youtube.com" in u or "youtu.be" in u or "vimeo.com" in u:
        return "video"

    if re.search(r"\.(mp3|wav|m4a|ogg|webm)(\?|$)", u):
        return "audio"
    if re.search(r"\.(mp4|mov|m4v|webm)(\?|$)", u):
        return "video"
    if re.search(r"\.(pdf)(\?|$)", u):
        return "pdf"
    if re.search(r"\.(png|jpg|jpeg|gif|webp)(\?|$)", u):
        return "image"

    if content_type_fallback in ("app", "website"):
        return content_type_fallback
    return "website"


def _guess_file_media_type(path: str, content_type_fallback: str) -> str:
    p = _norm(path).lower()
    if re.search(r"\.(mp3|wav|m4a|ogg|webm)$", p):
        return "audio"
    if re.search(r"\.(mp4|mov|m4v|webm)$", p):
        return "video"
    if p.endswith(".pdf"):
        return "pdf"
    if re.search(r"\.(png|jpg|jpeg|gif|webp)$", p):
        return "image"
    return content_type_fallback or "text"


def _is_raw_file_name(s: str) -> bool:
    """
    In the Excel file, 'Link' sometimes contains a filename like 'myfile.pdf'
    rather than a URL. Treat those as file-path references.
    """
    v = _norm(s)
    if not v:
        return False
    if v.startswith("http://") or v.startswith("https://"):
        return False
    return bool(re.search(r"\.(pdf|mp4|mov|m4a|mp3|wav|webm|ogg|png|jpg|jpeg|webp)$", v.lower()))


def _col_index_map(header_row: List[str]) -> Dict[str, int]:
    def find_col(pattern: str) -> Optional[int]:
        rx = re.compile(pattern, re.IGNORECASE)
        for i, h in enumerate(header_row):
            hh = _norm(h).strip()
            if rx.search(hh):
                return i
        return None

    return {
        "intervention_id": find_col(r"\bintervention[_\s-]*id\b|\bid\b"),
        "provider": find_col(r"\bprovider\b|\banbieter\b|\bquelle\b"),
        "link": find_col(r"^\s*link\s*$|\burl\b|\bweb\s*link\b|\bwebsite\b"),
        "title": find_col(r"^\s*title\s*$|\btitel\b|\bname\b"),
        "description": find_col(r"\bdescription\b|\bbeschreibung\b|\bdesc\b"),
        # This is the "Original language" column in your sheet (metadata)
        "language": find_col(r"\boriginal\s*language\b|\bsprache\b|\blanguage\b"),
        "aim": find_col(r"\baim\b|\bziel\b"),
        "topic": find_col(r"\btopic\b|\bthema\b"),
        # ✅ make content type very tolerant:
        "content_type": find_col(
            r"\bcontent\s*type\b|\bcontent[-_\s]*type\b|\bformat\b|\btype\b|\bmedien[-_\s]*typ\b|\binhalts[-_\s]*typ\b"
        ),
        "duration": find_col(r"\bduration\b|\bdauer\b|\bmin\b|\bminutes\b"),
        "lc9": find_col(r"\blc9\b"),
        "where": find_col(r"^\s*where\s*$|\bwo\b|\bord\b"),
        "setting": find_col(r"^\s*setting\s*$|\bumgebung\b|\bkontext\b"),
        "keywords": find_col(r"\bkeywords\b|\bstichw(ö|o)rter\b|\btags?\b"),
        # Optional extra fields if your sheet has them (won't break if missing)
        "input_from": find_col(r"\binput[_\s-]*from\b|\beingabe\b"),
        "primary_diagnosis": find_col(r"\bprimary\s*diagnosis\b|\bhauptdiagnose\b"),
        "original_language": find_col(r"\boriginal\s*language\b|\bsprache\b|\blanguage\b"),
        "cognitive_level": find_col(r"\bcognitive\s*level\b|\bkognitiv\b"),
        "physical_level": find_col(r"\bphysical\s*level\b|\bphysisch\b"),
        "frequency_time": find_col(r"\bfrequency\b|\bh(ä|a)ufigkeit\b"),
        "timing": find_col(r"\btiming\b|\bzeitpunkt\b"),
        "duration_bucket": find_col(r"\bduration\s*bucket\b|\bdauer\s*klasse\b"),
        "sex_specific": find_col(r"\bsex\s*specific\b|\bgeschlecht\b"),
    }


# ---------------- main import service ----------------


def import_interventions_from_excel(
    xlsm_path: str,
    sheet_name: str = "Content",
    dry_run: bool = False,
    limit: Optional[int] = None,
    default_lang: str = "en",
) -> Dict[str, Any]:
    """
    Imports interventions from Excel.

    Fixes:
    - Robustly finds columns even if headers contain suffixes like "(multi-choice)" / "(text input)"
      (this was why `where`, `setting`, and `media` stayed empty in your DB).
    - Splits multi-choice values not only by comma/semicolon but also by newline and "|".
    - Builds media[] from link column reliably.

    - Reads intervention_id like '4001_de'
    - Saves external_id='4001', language='de'
    - Upserts by (external_id, language)
    """
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=True)

    # ---- pick sheet (allow "Content (2)" etc.) ----
    chosen_sheet = None
    if sheet_name in wb.sheetnames:
        chosen_sheet = sheet_name
    else:
        target = _norm(sheet_name).lower()
        for sn in wb.sheetnames:
            if _norm(sn).lower().startswith(target):
                chosen_sheet = sn
                break
        if not chosen_sheet:
            for sn in wb.sheetnames:
                if "content" in _norm(sn).lower():
                    chosen_sheet = sn
                    break
    if not chosen_sheet:
        raise ValueError(f"Sheet '{sheet_name}' not found. Sheets: {wb.sheetnames}")

    ws = wb[chosen_sheet]

    header = [_norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    col = _col_index_map(header)

    # ---- header fallback: match substrings inside "where (multi-choice)" etc. ----
    def _find_header_fallback(patterns: List[str]) -> Optional[int]:
        for i, h in enumerate(header):
            hh = _norm(h).lower()
            for pat in patterns:
                if re.search(pat, hh, flags=re.IGNORECASE):
                    return i
        return None

    # strengthen required-ish + the ones that were missing for you
    if col.get("intervention_id") is None:
        col["intervention_id"] = _find_header_fallback([r"\bintervention[_\s-]*id\b", r"^\s*id\b"])
    if col.get("title") is None:
        col["title"] = _find_header_fallback([r"\btitle\b", r"\btitel\b", r"\bname\b"])
    if col.get("description") is None:
        col["description"] = _find_header_fallback([r"\bdescription\b", r"\bbeschreibung\b", r"\bdesc\b"])
    if col.get("content_type") is None:
        col["content_type"] = _find_header_fallback(
            [
                r"\bcontent\s*type\b",
                r"\bcontenttype\b",
                r"\bformat\b",
                r"\bmedium\b",
                r"\bmedien\b",
                r"\binhalts?\b",
            ]
        )

    # IMPORTANT FIXES (these headers are "where (multi-choice)" / "setting (multi-choice)" / "link (text input)")
    if col.get("where") is None:
        col["where"] = _find_header_fallback([r"\bwhere\b", r"\bwo\b", r"\bord\b"])
    if col.get("setting") is None:
        col["setting"] = _find_header_fallback([r"\bsetting\b", r"\bumgebung\b", r"\bkontext\b"])
    if col.get("link") is None:
        col["link"] = _find_header_fallback([r"\blink\b", r"\burl\b", r"\bweb\s*link\b", r"\bwebsite\b"])

    required = ["intervention_id", "title", "description", "content_type"]
    for r in required:
        if col.get(r) is None:
            raise ValueError(f"Missing required column mapping for: {r}")

    # local split that also handles newlines / pipes (typical REDCap exports)
    def _split_any(v: Any) -> List[str]:
        s = _norm(v)
        if not s:
            return []
        parts = re.split(r"[;,|\n]\s*", s)  # <-- key change vs old _split_list
        out: List[str] = []
        seen = set()
        for p in parts:
            p = _norm(p)
            if not p:
                continue
            k = p.lower()
            if k not in seen:
                out.append(p)
                seen.add(k)
        return out

    created = updated = skipped = 0
    errors: List[Dict[str, Any]] = []

    def _set_if_exists(doc: Any, field: str, value: Any) -> None:
        if hasattr(doc, field):
            setattr(doc, field, value)

    for row_idx in range(2, ws.max_row + 1):
        if limit and (created + updated + skipped) >= limit:
            break

        row = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        intervention_id_raw = _norm(row[col["intervention_id"]])
        if not intervention_id_raw:
            skipped += 1
            continue

        try:
            external_id, lang_from_id = _parse_external_id_and_language(intervention_id_raw)

            lang_from_col = _normalize_lang(row[col["language"]]) if col.get("language") is not None else None
            language = (lang_from_id or lang_from_col or default_lang).lower()

            provider = _norm(row[col["provider"]]) if col.get("provider") is not None else ""
            link_val = _norm(row[col["link"]]) if col.get("link") is not None else ""
            title = _norm(row[col["title"]])
            desc = _norm(row[col["description"]])

            aim = _norm(row[col["aim"]]) if col.get("aim") is not None else ""
            topic_raw = _norm(row[col["topic"]]) if col.get("topic") is not None else ""
            lc9_raw = _norm(row[col["lc9"]]) if col.get("lc9") is not None else ""
            where_raw = _norm(row[col["where"]]) if col.get("where") is not None else ""
            setting_raw = _norm(row[col["setting"]]) if col.get("setting") is not None else ""
            keywords_raw = _norm(row[col["keywords"]]) if col.get("keywords") is not None else ""

            content_type_raw = _norm(row[col["content_type"]])
            duration_raw = _norm(row[col["duration"]]) if col.get("duration") is not None else ""

            mapped_ct = _map_content_type(content_type_raw)
            duration_min = _parse_duration_minutes(duration_raw)

            # Lists (store as fields)
            topic_list = _split_any(topic_raw)
            lc9_list = _split_any(lc9_raw)
            where_list = _split_any(where_raw)
            setting_list = _split_any(setting_raw)
            keywords_list = _split_any(keywords_raw)

            existing = Intervention.objects(external_id=external_id, language=language).first()
            doc = existing or Intervention(external_id=external_id, language=language)

            _set_if_exists(doc, "provider", provider or getattr(doc, "provider", None))
            _set_if_exists(doc, "title", title or external_id)
            _set_if_exists(doc, "description", desc or "-")
            _set_if_exists(doc, "content_type", mapped_ct)

            if duration_min is not None:
                _set_if_exists(doc, "duration", duration_min)

            # store simple metadata if model has them
            if aim:
                _set_if_exists(doc, "aim", aim)

            orig_lang_val = lang_from_col or lang_from_id or language
            _set_if_exists(doc, "original_language", orig_lang_val)

            if topic_list:
                _set_if_exists(doc, "topic", topic_list)
            if lc9_list:
                _set_if_exists(doc, "lc9", lc9_list)
            if where_list:
                _set_if_exists(doc, "where", where_list)  # <-- should now fill: ["outside"]
            if setting_list:
                _set_if_exists(doc, "setting", setting_list)  # <-- should now fill: ["individual"]
            if keywords_list:
                _set_if_exists(doc, "keywords", keywords_list)

            if duration_raw:
                _set_if_exists(doc, "duration_bucket", duration_raw)

            # ---- media from link ----
            media_items: List[InterventionMedia] = []

            if link_val:
                if _is_raw_file_name(link_val):
                    media_items.append(
                        InterventionMedia(
                            kind="file",
                            media_type=_guess_file_media_type(link_val, mapped_ct.lower() if mapped_ct else "text"),
                            provider=None,
                            title=title or None,
                            file_path=link_val,
                            mime=None,
                        )
                    )
                else:
                    prov = _guess_provider(link_val)
                    embed = (
                        _spotify_embed(link_val)
                        if prov == "spotify"
                        else (_youtube_embed(link_val) if prov == "youtube" else None)
                    )
                    media_items.append(
                        InterventionMedia(
                            kind="external",
                            media_type=_guess_media_type_for_url(
                                link_val, mapped_ct.lower() if mapped_ct else "website"
                            ),
                            provider=prov,
                            title=title or None,
                            url=link_val,  # <-- correct: keep URL in url
                            embed_url=embed,
                        )
                    )

            def _media_key(m: InterventionMedia) -> str:
                return (
                    f"{_norm(getattr(m, 'kind', ''))}|"
                    f"{_norm(getattr(m, 'media_type', ''))}|"
                    f"{_norm(getattr(m, 'url', ''))}|"
                    f"{_norm(getattr(m, 'file_path', ''))}"
                )

            merged: List[InterventionMedia] = []
            seen_keys = set()

            for m in getattr(doc, "media", None) or []:
                k = _media_key(m)
                if k not in seen_keys:
                    merged.append(m)
                    seen_keys.add(k)

            for m in media_items:
                k = _media_key(m)
                if k not in seen_keys:
                    merged.append(m)
                    seen_keys.add(k)

            # Only overwrite media if link exists; otherwise keep previous media.
            if link_val:
                _set_if_exists(doc, "media", merged)  # <-- should now fill media for Vimeo/YouTube/etc.

            if not dry_run:
                doc.save()

            if existing:
                updated += 1
            else:
                created += 1

        except Exception as e:
            errors.append(
                {
                    "row": row_idx,
                    "intervention_id": intervention_id_raw,
                    "error": str(e),
                }
            )
            skipped += 1

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }
